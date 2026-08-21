"""
Professor-Student Matching System
Deferred Acceptance — รองรับทั้ง Student-Proposing และ Professor-Proposing
(v4 — รันทั้งสองแบบพร้อมกัน แสดงผลเปรียบเทียบสำหรับเสนอที่ประชุมคณะ)

อ่านข้อมูลดิบที่เก็บได้จาก MS Forms (ไฟล์ 04_raw_data_collected.xlsx) แล้วรัน
Gale-Shapley ทั้ง 2 แบบ โดยฝั่งอาจารย์รับได้หลายกลุ่มตาม Quota:

  - Student-Proposing: นักศึกษาเป็นฝั่งเสนอ -> Proposer-Optimal สำหรับนักศึกษา
  - Professor-Proposing: อาจารย์เป็นฝั่งเสนอ -> Proposer-Optimal สำหรับอาจารย์

ทั้งสองแบบยัง Stable เสมอ (ไม่มี Blocking Pair) ตามทฤษฎี Gale-Shapley เพียงแต่
"ฝั่งไหนเสนอ ฝั่งนั้นได้เปรียบ" (Roth, 1982) — สคริปต์นี้จึงรันทั้งคู่แล้วสรุปผล
เปรียบเทียบให้แอดมินนำไปเสนอที่ประชุมคณะเพื่อตัดสินใจเลือกแบบที่เหมาะกับนโยบาย

ขั้นตอน tie-break ของฝั่งอาจารย์ (ตามนโยบายที่ตกลงกันไว้ล่าสุด) ใช้เพื่อแปลง
Main Score (1-100) ให้กลายเป็น "Strict Preference List" ก่อนป้อนเข้า Algorithm
มาตรฐาน — ใช้ Preference List เดียวกันนี้กับทั้งสองแบบ (Tie-break ไม่ได้ขึ้นกับ
ว่าใครเป็นฝั่งเสนอ) วิธีนี้ให้ผลลัพธ์เทียบเท่ากับการรัน Extended Gale-Shapley
สำหรับ One-sided Ties (Irving) และยัง Audit ได้ทุกขั้นตอน:

    1) Main Score (1-100, ปัดเศษจาก 0.5A+0.5B) — มากไปน้อย
    2) SubScore แบบ decimal (ไม่ปัดเศษ)         — มากไปน้อย
    3) อันดับที่กลุ่มนั้น Rank อาจารย์ท่านนี้ไว้   — น้อยไปมาก (rank 1 ดีที่สุด)
    4) Seeded Random (บันทึก seed ไว้ตรวจสอบย้อนหลังได้)

เงื่อนไขที่ต้อง Enforce ก่อนรัน (ตามที่ตกลงกันไว้):
    - ฝั่งนักศึกษาต้อง Rank ครบทุกคน (Complete Strict List) ห้ามเว้น/ซ้ำ
    - ผลรวม Quota ฝั่งอาจารย์ต้อง >= จำนวนกลุ่มนักศึกษาทั้งหมด

Usage:
    python 05_matching_algorithm.py --input 04_raw_data_collected.xlsx \
                                     --output 06_matching_comparison.xlsx \
                                     --seed 2026

    # รันแค่แบบเดียวก็ได้ ถ้าไม่ต้องการเปรียบเทียบ:
    python 05_matching_algorithm.py --mode student
    python 05_matching_algorithm.py --mode professor
"""

import argparse
import random
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="D9E2F3")
OK_FILL = PatternFill("solid", fgColor="E2F0D9")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# --------------------------------------------------------------------------
# 1) LOAD DATA
# --------------------------------------------------------------------------
def load_data(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    # Professor_Info -> quotas
    ws = wb["Professor_Info"]
    quotas = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue
        # Support both 4-column [ProfID, FullName, Expertise, Quota]
        # and 5-column [ProfID, AnonymousCode, FullName, Expertise, Quota]
        if len(row) >= 5 and row[1] is not None and str(row[1]).strip().startswith("P"):
            code = str(row[1]).strip()
            quota = row[4]
        else:
            code = str(row[0]).strip()
            quota = row[3] if len(row) >= 4 else row[-1]
        if code is None or quota is None:
            continue
        quotas[code] = int(quota)

    # Student_Rankings -> group -> {prof_code: rank}
    ws = wb["Student_Rankings"]
    prof_codes = [str(c.value).strip() for c in ws[1][1:] if c.value is not None]
    student_pref = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        gcode = str(row[0]).strip()
        ranks = dict(zip(prof_codes, row[1:]))
        # sort professors by rank ascending (1 = most preferred) -> ordered list
        ordered = [p for p, _ in sorted(ranks.items(), key=lambda kv: kv[1])]
        student_pref[gcode] = {"order": ordered, "rank_of": ranks}

    # Professor_Scores -> (prof, group) -> dict of scores
    ws = wb["Professor_Scores"]
    prof_scores = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        pcode, gcode, a, b, subscore, mainscore = row[:6]
        pcode_str = str(pcode).strip()
        gcode_str = str(gcode).strip()
        prof_scores[pcode_str][gcode_str] = {
            "A": a, "B": b,
            "sub": float(subscore), "main": int(round(mainscore)),
        }

    groups = list(student_pref.keys())
    profs = list(quotas.keys())
    return groups, profs, quotas, student_pref, prof_scores


# --------------------------------------------------------------------------
# 2) BUILD STRICT PROFESSOR PREFERENCE LISTS (apply tie-break chain)
# --------------------------------------------------------------------------
def build_prof_preferences(profs, groups, prof_scores, student_pref, seed):
    rng = random.Random(seed)
    prof_pref = {}
    tie_log = []

    for p in profs:
        scored = []
        for g in groups:
            s = prof_scores[p][g]
            rank_by_group = student_pref[g]["rank_of"][p]  # lower = more preferred by group
            scored.append((g, s["main"], s["sub"], rank_by_group))

        # detect ties at (main, sub, rank_by_group) level -> need random tiebreak
        random_keys = {}
        seen_keys = defaultdict(list)
        for g, main, sub, rbg in scored:
            key = (main, round(sub, 6), rbg)
            seen_keys[key].append(g)
        for key, gs in seen_keys.items():
            if len(gs) > 1:
                shuffled = gs[:]
                rng.shuffle(shuffled)
                for i, g in enumerate(shuffled):
                    random_keys[g] = i
                tie_log.append({
                    "prof": p, "groups": ",".join(gs),
                    "main": key[0], "sub": key[1], "rank_by_group": key[2],
                    "resolved_order": ",".join(shuffled),
                })
            else:
                random_keys[gs[0]] = 0

        # Sort: main desc, sub desc, rank_by_group asc, then random_key asc
        scored.sort(key=lambda t: (-t[1], -t[2], t[3], random_keys[t[0]]))
        prof_pref[p] = [g for g, *_ in scored]

    return prof_pref, tie_log


# --------------------------------------------------------------------------
# 3a) STUDENT-PROPOSING DEFERRED ACCEPTANCE (Gale-Shapley with quotas)
#     นักศึกษาเสนอตัวไปหาอาจารย์ตามลำดับที่ตัวเองอยากได้ที่สุดก่อน
# --------------------------------------------------------------------------
def student_proposing_da(groups, profs, quotas, student_pref, prof_pref):
    free_groups = list(groups)
    next_proposal_idx = {g: 0 for g in groups}
    prof_holds = {p: [] for p in profs}  # currently held groups, best-first
    prof_rank_index = {p: {g: i for i, g in enumerate(prof_pref[p])} for p in profs}

    while free_groups:
        g = free_groups.pop(0)
        order = student_pref[g]["order"]
        idx = next_proposal_idx[g]
        if idx >= len(order):
            continue  # exhausted list (should not happen if list is complete)
        p = order[idx]
        next_proposal_idx[g] = idx + 1

        held = prof_holds[p]
        held.append(g)
        # keep only top-`quota` by professor's strict preference; reject the rest
        held.sort(key=lambda x: prof_rank_index[p][x])
        quota = quotas[p]
        if len(held) > quota:
            rejected = held[quota:]
            prof_holds[p] = held[:quota]
            for rg in rejected:
                free_groups.append(rg)
        else:
            prof_holds[p] = held

    matching = {}
    for p, gs in prof_holds.items():
        for g in gs:
            matching[g] = p
    unmatched = [g for g in groups if g not in matching]
    return matching, unmatched, prof_holds


# --------------------------------------------------------------------------
# 3b) PROFESSOR-PROPOSING DEFERRED ACCEPTANCE (Gale-Shapley with quotas)
#     อาจารย์เสนอตัวไปหากลุ่มนักศึกษาตามลำดับที่ตัวเองอยากได้ที่สุดก่อน
#     (แต่ละอาจารย์เสนอได้พร้อมกันหลายที่นั่งตาม Quota ที่ยังว่าง)
# --------------------------------------------------------------------------
def professor_proposing_da(groups, profs, quotas, student_pref, prof_pref):
    next_idx = {p: 0 for p in profs}
    remaining_quota = {p: quotas[p] for p in profs}
    held = {g: None for g in groups}  # group -> currently held professor (or None)
    student_rank_index = {g: {p: i for i, p in enumerate(student_pref[g]["order"])} for g in groups}

    free_profs = [p for p in profs if quotas[p] > 0]
    proposed_already = {p: set() for p in profs}

    while free_profs:
        p = free_profs.pop(0)
        while next_idx[p] < len(prof_pref[p]):
            g = prof_pref[p][next_idx[p]]
            next_idx[p] += 1
            if g in proposed_already[p]:
                continue
            proposed_already[p].add(g)

            current = held[g]
            if current is None:
                held[g] = p
                remaining_quota[p] -= 1
                break
            elif student_rank_index[g][p] < student_rank_index[g][current]:
                # กลุ่มนี้ชอบอาจารย์ p มากกว่าคนที่ถืออยู่ -> สลับ, คืน quota ให้อาจารย์เดิม
                held[g] = p
                remaining_quota[p] -= 1
                remaining_quota[current] += 1
                if current not in free_profs and remaining_quota[current] > 0 \
                        and next_idx[current] < len(prof_pref[current]):
                    free_profs.append(current)
                break
            # else: กลุ่มปฏิเสธ p (ชอบคนที่ถืออยู่มากกว่า) -> ลองเสนอกลุ่มถัดไป
        if remaining_quota[p] > 0 and next_idx[p] < len(prof_pref[p]):
            free_profs.append(p)

    matching = {g: p for g, p in held.items() if p is not None}
    unmatched = [g for g in groups if g not in matching]
    prof_holds = defaultdict(list)
    for g, p in matching.items():
        prof_holds[p].append(g)
    return matching, unmatched, dict(prof_holds)


# --------------------------------------------------------------------------
# 4) HELPERS FOR STATS
# --------------------------------------------------------------------------
def compute_stats(groups, profs, student_pref, prof_scores, matching, unmatched):
    n = len(groups)
    matched_ranks = [student_pref[g]["rank_of"][matching[g]] for g in groups if g in matching]
    matched_mains = [prof_scores[matching[g]][g]["main"] for g in groups if g in matching]
    stats = {
        "n_groups": n,
        "n_matched": len(matched_ranks),
        "n_unmatched": len(unmatched),
        "avg_rank": round(sum(matched_ranks) / len(matched_ranks), 2) if matched_ranks else None,
        "pct_top1": round(100 * sum(1 for r in matched_ranks if r == 1) / n, 1) if matched_ranks else None,
        "pct_top3": round(100 * sum(1 for r in matched_ranks if r <= 3) / n, 1) if matched_ranks else None,
        "avg_prof_main_score": round(sum(matched_mains) / len(matched_mains), 2) if matched_mains else None,
    }
    return stats


# --------------------------------------------------------------------------
# 5) WRITE ONE MATCHING RESULT INTO A GIVEN WORKBOOK (as a labeled sheet group)
# --------------------------------------------------------------------------
def write_mode_sheets(wb, label, groups, profs, quotas, student_pref, prof_scores,
                       matching, unmatched, prof_holds):
    # ---- Final_Matching_<label> ----
    ws = wb.create_sheet(f"Final_Matching_{label}")
    headers = ["GroupCode", "AssignedProfessor", "RankGroupGaveProf(1=best)",
               "MainScoreProfGaveGroup", "SubScoreDecimal"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for g in groups:
        p = matching.get(g)
        if p is None:
            ws.append([g, "UNMATCHED", None, None, None])
            continue
        rank_given = student_pref[g]["rank_of"][p]
        s = prof_scores[p][g]
        ws.append([g, p, rank_given, s["main"], round(s["sub"], 3)])
    autosize(ws, [12, 16, 22, 20, 16])
    ws.freeze_panes = "A2"

    # ---- Professor_Summary_<label> ----
    ws = wb.create_sheet(f"Professor_Summary_{label}")
    headers = ["ProfCode", "Quota", "GroupsAssigned", "NumAssigned", "QuotaRemaining"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for p in profs:
        assigned = prof_holds.get(p, [])
        ws.append([p, quotas[p], ", ".join(assigned) if assigned else "-",
                   len(assigned), quotas[p] - len(assigned)])
    autosize(ws, [10, 8, 40, 14, 16])


def write_tie_log_sheet(wb, tie_log):
    ws = wb.create_sheet("TieBreak_Log")
    headers = ["ProfCode", "GroupsTied", "MainScore", "SubScore", "RankByGroup", "ResolvedOrder(bestFirst)"]
    ws.append(headers)
    style_header(ws, 1, len(headers))
    for entry in tie_log:
        ws.append([entry["prof"], entry["groups"], entry["main"], entry["sub"],
                   entry["rank_by_group"], entry["resolved_order"]])
    autosize(ws, [10, 24, 12, 12, 14, 30])
    if not tie_log:
        ws.append(["-", "ไม่มี Tie ที่ต้องใช้ Seeded Random ในรอบนี้", "", "", "", ""])
    note_note = "หมายเหตุ: Tie-break policy ใช้ชุดเดียวกันทั้ง Student-Proposing และ Professor-Proposing (ไม่ขึ้นกับฝั่งที่เสนอ)"
    ws.append([])
    ws.append([note_note])


def write_comparison_sheet(wb, stats_student, stats_prof, seed):
    ws = wb.create_sheet("Comparison_Summary")
    ws.sheet_view.showGridLines = True
    headers = ["ตัวชี้วัด", "Student-Proposing", "Professor-Proposing", "ฝั่งที่ได้เปรียบ"]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    def better(a, b, higher_is_better=True):
        if a is None or b is None:
            return "-"
        if a == b:
            return "เท่ากัน"
        if higher_is_better:
            return "Student" if a > b else "Professor"
        return "Student" if a < b else "Professor"

    rows = [
        ("จำนวนกลุ่มที่จับคู่สำเร็จ", stats_student["n_matched"], stats_prof["n_matched"],
         better(stats_student["n_matched"], stats_prof["n_matched"])),
        ("จำนวนกลุ่มที่ไม่ได้จับคู่ (Unmatched)", stats_student["n_unmatched"], stats_prof["n_unmatched"],
         better(stats_student["n_unmatched"], stats_prof["n_unmatched"], higher_is_better=False)),
        ("Rank เฉลี่ยที่นักศึกษาได้รับ (ยิ่งน้อยยิ่งดี)", stats_student["avg_rank"], stats_prof["avg_rank"],
         better(stats_student["avg_rank"], stats_prof["avg_rank"], higher_is_better=False)),
        ("% กลุ่มที่ได้อาจารย์อันดับ 1 ของตน", stats_student["pct_top1"], stats_prof["pct_top1"],
         better(stats_student["pct_top1"], stats_prof["pct_top1"])),
        ("% กลุ่มที่ได้อาจารย์อยู่ใน Top-3 ของตน", stats_student["pct_top3"], stats_prof["pct_top3"],
         better(stats_student["pct_top3"], stats_prof["pct_top3"])),
        ("Main Score เฉลี่ยที่อาจารย์ให้กลุ่มที่ตนได้ (ความพอใจฝั่งอาจารย์)",
         stats_student["avg_prof_main_score"], stats_prof["avg_prof_main_score"],
         better(stats_student["avg_prof_main_score"], stats_prof["avg_prof_main_score"])),
    ]
    for r in rows:
        ws.append(list(r))
    autosize(ws, [50, 20, 20, 18])

    ws.append([])
    ws.append(["Random Seed ที่ใช้ตัดสิน Tie (สำหรับตรวจสอบย้อนหลัง)", seed])
    ws.append([])
    ws.append(["คำอธิบาย (สำหรับนำเสนอที่ประชุมคณะ)"])
    explain = (
        "ทั้งสองแบบเป็น Stable Matching เสมอ (ไม่มี Blocking Pair) ตามทฤษฎี Gale-Shapley "
        "แต่ 'ฝั่งไหนเสนอ ฝั่งนั้นได้เปรียบ' (Roth, 1982): Student-Proposing ให้ผลลัพธ์ที่ดีที่สุด "
        "เท่าที่เป็นไปได้จากมุมนักศึกษา (Proposer-Optimal) ส่วน Professor-Proposing ให้ผลลัพธ์ที่ดี"
        "ที่สุดจากมุมอาจารย์แทน — ไม่มีแบบไหน 'ดีกว่า' แบบเบ็ดเสร็จ ขึ้นอยู่กับนโยบายที่คณะต้องการให้"
        "น้ำหนักกับฝั่งใดมากกว่า"
    )
    ws.append([explain])
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=4)
    ws.cell(row=ws.max_row, column=1).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[ws.max_row].height = 60


# --------------------------------------------------------------------------
# 6) MAIN WRITE ENTRYPOINT
# --------------------------------------------------------------------------
def write_output(path, groups, profs, quotas, student_pref, prof_scores, tie_log, seed,
                  result_student=None, result_prof=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default blank sheet; we add named sheets explicitly

    stats_student = stats_prof = None

    if result_student is not None:
        matching, unmatched, prof_holds = result_student
        write_mode_sheets(wb, "Student", groups, profs, quotas, student_pref,
                           prof_scores, matching, unmatched, prof_holds)
        stats_student = compute_stats(groups, profs, student_pref, prof_scores, matching, unmatched)

    if result_prof is not None:
        matching, unmatched, prof_holds = result_prof
        write_mode_sheets(wb, "Professor", groups, profs, quotas, student_pref,
                           prof_scores, matching, unmatched, prof_holds)
        stats_prof = compute_stats(groups, profs, student_pref, prof_scores, matching, unmatched)

    if stats_student is not None and stats_prof is not None:
        write_comparison_sheet(wb, stats_student, stats_prof, seed)

    write_tie_log_sheet(wb, tie_log)
    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="04_raw_data_collected.xlsx")
    ap.add_argument("--output", default="06_matching_comparison.xlsx")
    ap.add_argument("--seed", type=int, default=2026,
                     help="Seed สำหรับ Tie-break ชั้นสุดท้าย (บันทึกไว้เพื่อ Audit)")
    ap.add_argument("--mode", choices=["student", "professor", "both"], default="both",
                     help="รันแบบไหน: student-proposing, professor-proposing, หรือทั้งคู่ (default)")
    args = ap.parse_args()

    groups, profs, quotas, student_pref, prof_scores = load_data(args.input)

    total_quota = sum(quotas.values())
    if total_quota < len(groups):
        raise SystemExit(
            f"[ERROR] ผลรวม Quota ({total_quota}) น้อยกว่าจำนวนกลุ่มนักศึกษา ({len(groups)}) "
            f"— ต้องเพิ่ม Quota ก่อนรัน Matching"
        )

    for g in groups:
        order = student_pref[g]["order"]
        if len(order) != len(profs) or len(set(order)) != len(profs):
            raise SystemExit(f"[ERROR] กลุ่ม {g} จัดอันดับอาจารย์ไม่ครบ/ไม่ถูกต้อง (ต้อง Rank ครบ {len(profs)} ท่าน แบบไม่ซ้ำ)")

    # Tie-break chain ใช้ Preference List ชุดเดียวกันสำหรับทั้งสองแบบ (ไม่ขึ้นกับฝั่งที่เสนอ)
    prof_pref, tie_log = build_prof_preferences(profs, groups, prof_scores, student_pref, args.seed)

    result_student = result_prof = None

    if args.mode in ("student", "both"):
        result_student = student_proposing_da(groups, profs, quotas, student_pref, prof_pref)
        m, u, _ = result_student
        print(f"[Student-Proposing]   Matched: {len(m)}/{len(groups)} | Unmatched: {u}")

    if args.mode in ("professor", "both"):
        result_prof = professor_proposing_da(groups, profs, quotas, student_pref, prof_pref)
        m, u, _ = result_prof
        print(f"[Professor-Proposing] Matched: {len(m)}/{len(groups)} | Unmatched: {u}")

    write_output(args.output, groups, profs, quotas, student_pref, prof_scores, tie_log, args.seed,
                 result_student=result_student, result_prof=result_prof)

    print(f"Tie-break events resolved with seed={args.seed}: {len(tie_log)}")
    print(f"Output written to: {args.output}")
    if args.mode == "both":
        print("-> ดูสรุปเปรียบเทียบได้ที่ชีต 'Comparison_Summary' ในไฟล์ผลลัพธ์")


if __name__ == "__main__":
    main()
