"""
Matching Job Service
Wraps 05_matching_algorithm.py as a backend job
"""
import os
import sys
import json
import re
import tempfile
from typing import Optional
from datetime import datetime
from sqlalchemy.orm import Session
import openpyxl
import models

# Path to the matching algorithm script
ALGORITHM_PATH = os.path.join(
    os.path.dirname(__file__), "..", "..", "materials", "05_matching_algorithm.py"
)
RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "data", "results")


def _export_input_excel(session_id: int, program: Optional[str], db: Session, tmp_path: str):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, size=11, color="1F2937")
    header_fill = PatternFill("solid", fgColor="E6F4FF")
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB")
    )
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def style_sheet_header(ws, col_count: int):
        ws.row_dimensions[1].height = 28
        for col in range(1, col_count + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center
            cell.border = thin_border

    def auto_fit_columns(ws, min_widths=None):
        if min_widths is None:
            min_widths = {}
        for col in ws.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                lines = val.split('\n')
                for line in lines:
                    max_len = max(max_len, len(line))
            width = max(max_len + 4, min_widths.get(col_letter, 12))
            ws.column_dimensions[col_letter].width = min(width, 50)

    q_groups = db.query(models.Group).filter_by(session_id=session_id)
    q_profs = db.query(models.Professor).filter_by(session_id=session_id)
    q_rankings = db.query(models.StudentRanking).filter_by(session_id=session_id)
    q_scores = db.query(models.ProfessorScore).filter_by(session_id=session_id)

    if program:
        has_prog = db.query(models.Group).filter(models.Group.session_id == session_id, models.Group.program != None).count() > 0
        if has_prog:
            q_groups = q_groups.filter_by(program=program)
            q_profs = q_profs.filter_by(program=program)
            q_rankings = q_rankings.filter_by(program=program)
            q_scores = q_scores.filter_by(program=program)

    groups = q_groups.all()
    professors = q_profs.all()
    rankings = q_rankings.all()
    scores = q_scores.all()

    # Sort groups by GroupID ascending
    def get_group_sort_key(g):
        code = str(g.anonymous_code or g.group_id or g.id)
        num_part = re.findall(r'\d+', code)
        if num_part:
            return (0, int(num_part[0]), code)
        return (1, 0, code)

    groups = sorted(groups, key=get_group_sort_key)

    # Sort professors by ProfID ascending
    def get_prof_sort_key(p):
        code = str(p.anonymous_code or p.prof_id or p.id)
        num_part = re.findall(r'\d+', code)
        if num_part:
            return (0, int(num_part[0]), code)
        return (1, 0, code)

    professors = sorted(professors, key=get_prof_sort_key)

    wb = openpyxl.Workbook()

    # ── 1. Group_Info ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Group_Info"
    group_headers = [
        "GroupID",
        "จำนวนสมาชิกกลุ่ม",
        "รหัสนักศึกษาของสมาชิก",
        "ชื่อ-นามสกุลของสมาชิก",
        "หัวข้อที่สนใจ",
        "รายละเอียดของหัวข้อ",
    ]
    ws.append(group_headers)
    style_sheet_header(ws, len(group_headers))

    for row_idx, g in enumerate(groups, start=2):
        group_id_val = g.anonymous_code or g.group_id or f"G{row_idx-1:03d}"
        
        student_ids = []
        student_names = []
        if g.members:
            for m in g.members:
                if m.student_id:
                    student_ids.append(str(m.student_id).strip())
                if m.full_name:
                    student_names.append(str(m.full_name).strip())
        
        ids_str = "\n".join(student_ids) if student_ids else (str(g.group_id or g.anonymous_code or "—"))
        names_str = "\n".join(student_names) if student_names else (str(g.representative or "—"))
        member_count = g.member_count or (len(g.members) if g.members else (len(student_ids) if student_ids else 1))

        topics = []
        try:
            topics = json.loads(g.topic_interest) if g.topic_interest else []
        except Exception:
            pass
        
        topic_titles = []
        topic_details = []
        if isinstance(topics, list):
            for i, t in enumerate(topics, start=1):
                if isinstance(t, dict):
                    title = t.get("title", "").strip()
                    detail = t.get("detail", "").strip()
                    if title:
                        topic_titles.append(f"{i}. {title}")
                    if detail and detail != "—":
                        topic_details.append(f"{i}. {detail}")
                    elif title:
                        topic_details.append(f"{i}. —")
                elif isinstance(t, str) and t.strip():
                    topic_titles.append(f"{i}. {t.strip()}")
                    topic_details.append(f"{i}. —")
        
        titles_str = "\n".join(topic_titles) if topic_titles else "—"
        details_str = "\n".join(topic_details) if topic_details else "—"

        row_data = [group_id_val, member_count, ids_str, names_str, titles_str, details_str]
        ws.append(row_data)

        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in [1, 2]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    auto_fit_columns(ws, {"A": 12, "B": 18, "C": 22, "D": 25, "E": 35, "F": 45})

    # ── 2. Professor_Info ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Professor_Info")
    prof_headers = [
        "ProfID",
        "ชื่อ-นามสกุลของอาจารย์",
        "ความเชี่ยวชาญ",
        "โควต้ากลุ่ม",
    ]
    ws2.append(prof_headers)
    style_sheet_header(ws2, len(prof_headers))

    for row_idx, p in enumerate(professors, start=2):
        prof_id_val = p.anonymous_code or p.prof_id or f"P{row_idx-1:03d}"
        row_data = [
            prof_id_val,
            p.full_name or "—",
            p.expertise or "—",
            p.quota if p.quota is not None else 0,
        ]
        ws2.append(row_data)

        for col_idx in range(1, len(row_data) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            if col_idx in [1, 4]:
                cell.alignment = align_center
            else:
                cell.alignment = align_left

    auto_fit_columns(ws2, {"A": 12, "B": 25, "C": 40, "D": 15})

    # ── 3. Student_Rankings (ไม่เปลี่ยนแปลงเนื้อหา) ────────────────────────
    ws3 = wb.create_sheet("Student_Rankings")
    prof_codes = [p.anonymous_code or p.prof_id for p in professors]
    ranking_headers = ["GroupID"] + prof_codes
    ws3.append(ranking_headers)
    style_sheet_header(ws3, len(ranking_headers))

    rank_lookup = {}
    for r in rankings:
        rank_lookup[(r.group_code, r.prof_code)] = r.rank

    for row_idx, g in enumerate(groups, start=2):
        gcode = g.anonymous_code or g.group_id
        row_data = [gcode] + [rank_lookup.get((gcode, pc)) for pc in prof_codes]
        ws3.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = align_center

    auto_fit_columns(ws3, {"A": 14})

    # ── 4. Professor_Scores (ไม่เปลี่ยนแปลงเนื้อหา) ────────────────────────
    ws4 = wb.create_sheet("Professor_Scores")
    scores_headers = [
        "ProfID",
        "GroupID",
        "Score_TopicFit_A",
        "Score_Clarity_B",
        "SubScore",
        "MainScore",
    ]
    ws4.append(scores_headers)
    style_sheet_header(ws4, len(scores_headers))

    sorted_scores = sorted(scores, key=lambda s: (str(s.prof_code), str(s.group_code)))
    for row_idx, s in enumerate(sorted_scores, start=2):
        row_data = [s.prof_code, s.group_code, s.score_a, s.score_b, s.sub_score, s.main_score]
        ws4.append(row_data)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws4.cell(row=row_idx, column=col_idx)
            cell.border = thin_border
            cell.alignment = align_center

    auto_fit_columns(ws4)

    wb.save(tmp_path)


def run_matching(run_id: int, session_id: int, seed: int, program: str, db: Session):
    """
    Execute the matching algorithm as a subprocess.
    Updates the MatchingRun record with results.
    """
    import subprocess
    os.makedirs(RESULTS_DIR, exist_ok=True)

    run = db.query(models.MatchingRun).filter_by(id=run_id).first()
    if not run:
        return

    # Temp input file
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp_input = f.name

    # Output file path
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    output_filename = f"matching_result_{timestamp}_seed{seed}.xlsx"
    output_path = os.path.join(RESULTS_DIR, output_filename)

    try:
        _export_input_excel(session_id, program, db, tmp_input)

        # Run the algorithm script
        result = subprocess.run(
            [sys.executable, os.path.abspath(ALGORITHM_PATH),
             "--input", tmp_input,
             "--output", output_path,
             "--seed", str(seed)],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )

        stdout = result.stdout if result.stdout is not None else ""
        stderr = result.stderr if result.stderr is not None else ""
        log_output = stdout + ("\nSTDERR:\n" + stderr if stderr else "")

        if result.returncode != 0:
            run.status = "failed"
            run.log = log_output
            db.commit()
            return

        # ── Parse results from stdout (v4 format) ──────────────────────────
        # Example lines:
        #   [Student-Proposing]   Matched: 10/15 | Unmatched: ['G001']
        #   [Professor-Proposing] Matched: 11/15 | Unmatched: []
        #   Tie-break events resolved with seed=2026: 3
        num_matched_student = 0
        num_unmatched_student = 0
        num_matched_professor = 0
        num_unmatched_professor = 0
        num_ties = 0

        for line in result.stdout.splitlines():
            # v4 format: [Student-Proposing]   Matched: X/Y | Unmatched: [...]
            m = re.match(r'\[Student-Proposing\].*Matched:\s*(\d+)/(\d+)', line)
            if m:
                num_matched_student = int(m.group(1))
                total = int(m.group(2))
                num_unmatched_student = total - num_matched_student
                continue

            m = re.match(r'\[Professor-Proposing\].*Matched:\s*(\d+)/(\d+)', line)
            if m:
                num_matched_professor = int(m.group(1))
                total = int(m.group(2))
                num_unmatched_professor = total - num_matched_professor
                continue

            if "Tie-break events" in line:
                parts = re.findall(r'\d+', line.split("seed=")[1] if "seed=" in line else line)
                # "Tie-break events resolved with seed=2026: 3"  → last number is count
                tie_parts = line.split(":")
                if len(tie_parts) >= 2:
                    try:
                        num_ties = int(tie_parts[-1].strip().split()[0])
                    except (ValueError, IndexError):
                        pass

        # Read output Excel and store results in DB (both modes)
        _import_results(run_id, output_path, db)

        # Use student-proposing as the primary "legacy" aggregate
        run.status = "success"
        run.num_matched = num_matched_student
        run.num_unmatched = num_unmatched_student
        run.num_ties = num_ties
        run.num_matched_student = num_matched_student
        run.num_unmatched_student = num_unmatched_student
        run.num_matched_professor = num_matched_professor
        run.num_unmatched_professor = num_unmatched_professor
        run.output_file_path = output_path
        run.log = log_output
        db.commit()

    except subprocess.TimeoutExpired:
        run.status = "failed"
        run.log = "Algorithm timed out after 120 seconds"
        db.commit()
    except Exception as e:
        run.status = "failed"
        run.log = str(e)
        db.commit()
    finally:
        if os.path.exists(tmp_input):
            os.unlink(tmp_input)


def _import_results(run_id: int, output_path: str, db: Session):
    """
    Read output Excel and store results in matching_results table.
    Supports both v4 (Final_Matching_Student / Final_Matching_Professor)
    and legacy v3 (Final_Matching) sheet naming.
    """
    wb = openpyxl.load_workbook(output_path, data_only=True)
    sheetnames = wb.sheetnames

    # v4: two mode sheets
    mode_map = {}
    if "Final_Matching_Student" in sheetnames:
        mode_map["student"] = wb["Final_Matching_Student"]
    if "Final_Matching_Professor" in sheetnames:
        mode_map["professor"] = wb["Final_Matching_Professor"]

    # Fallback to legacy v3 sheet name
    if not mode_map and "Final_Matching" in sheetnames:
        mode_map["student"] = wb["Final_Matching"]

    if not mode_map:
        return

    for mode, ws in mode_map.items():
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            group_code, assigned_prof, rank_given, main_score, sub_score = row
            result = models.MatchingResult(
                run_id=run_id,
                group_code=str(group_code),
                assigned_prof=str(assigned_prof) if assigned_prof else None,
                rank_given=int(rank_given) if rank_given is not None else None,
                main_score=int(main_score) if main_score is not None else None,
                sub_score=float(sub_score) if sub_score is not None else None,
                mode=mode,
            )
            db.add(result)
