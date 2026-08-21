"""
Excel Import Service
อ่านไฟล์ Excel 4 ชีต → บันทึกลง Database
"""
import json
from typing import Tuple, Optional
import openpyxl
from sqlalchemy.orm import Session
import models


def import_excel(file_path: str, filename: str, db: Session, program: Optional[str] = None) -> Tuple[int, str]:
    """
    Import Excel file with 4 sheets into DB.
    Returns (session_id, error_message_or_empty)
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        return -1, f"ไม่สามารถเปิดไฟล์ Excel ได้: {str(e)}"

    required_sheets = {"Group_Info", "Professor_Info", "Student_Rankings", "Professor_Scores"}
    missing = required_sheets - set(wb.sheetnames)
    if missing:
        return -1, f"ไฟล์ขาดชีต: {', '.join(missing)}"

    # Create session
    session = models.ImportSession(filename=filename, file_path=file_path, status="pending")
    db.add(session)
    db.flush()
    sid = session.id

    try:
        _import_group_info(wb["Group_Info"], sid, db, program)
        _import_professor_info(wb["Professor_Info"], sid, db, program)
        _import_student_rankings(wb["Student_Rankings"], sid, db, program)
        _import_professor_scores(wb["Professor_Scores"], sid, db, program)
        session.status = "imported"
        db.commit()
        return sid, ""
    except Exception as e:
        db.rollback()
        return -1, f"เกิดข้อผิดพลาดระหว่าง Import: {str(e)}"


def _import_group_info(ws, session_id: int, db: Session, program: Optional[str] = None):
    """
    ชีต Group_Info รองรับทั้ง:
    - รูปแบบใหม่ (6 คอลัมน์): GroupID | จำนวนสมาชิกกลุ่ม | รหัสนักศึกษาของสมาชิก | ชื่อ-นามสกุลของสมาชิก | หัวข้อที่สนใจ | รายละเอียดของหัวข้อ
    - รูปแบบเดิม (7 คอลัมน์): GroupID | AnonymousCode | Representative | MemberCount | Topic1 | Topic2 | Topic3
    """
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    header_row = [str(c).strip().lower() if c else "" for c in (first_row or [])]
    is_new_format = any("รหัส" in h or "สมาชิกกลุ่ม" in h or "รายละเอียด" in h for h in header_row) or len(header_row) == 6

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue

        group_id_str = str(row[0]).strip()

        if is_new_format:
            member_count = int(row[1]) if len(row) > 1 and row[1] is not None else None
            student_ids_raw = [s.strip() for s in str(row[2]).strip().splitlines() if s.strip()] if len(row) > 2 and row[2] is not None else []
            student_names_raw = [s.strip() for s in str(row[3]).strip().splitlines() if s.strip()] if len(row) > 3 and row[3] is not None else []
            topics_raw = [s.strip() for s in str(row[4]).strip().splitlines() if s.strip()] if len(row) > 4 and row[4] is not None else []
            details_raw = [s.strip() for s in str(row[5]).strip().splitlines() if s.strip()] if len(row) > 5 and row[5] is not None else []

            # Clean topic strings (remove leading "1. ", "2. ", etc.)
            topics_list = []
            max_topics = max(len(topics_raw), len(details_raw))
            for idx in range(max_topics):
                t_str = topics_raw[idx] if idx < len(topics_raw) else ""
                d_str = details_raw[idx] if idx < len(details_raw) else ""
                t_clean = re.sub(r'^\d+\.\s*', '', t_str).strip()
                d_clean = re.sub(r'^\d+\.\s*', '', d_str).strip()
                if t_clean or d_clean:
                    entry = {"title": t_clean}
                    if d_clean and d_clean != "—":
                        entry["detail"] = d_clean
                    topics_list.append(entry)

            rep_name = student_names_raw[0] if student_names_raw else None

            group = models.Group(
                session_id=session_id,
                group_id=group_id_str,
                anonymous_code=group_id_str,
                program=program,
                representative=rep_name,
                member_count=member_count or len(student_ids_raw) or 1,
                topic_interest=json.dumps(topics_list, ensure_ascii=False),
            )
            db.add(group)
            db.flush()

            max_members = max(len(student_ids_raw), len(student_names_raw))
            for idx in range(max_members):
                sid = student_ids_raw[idx] if idx < len(student_ids_raw) else f"STD-{group.id}-{idx+1}"
                sname = student_names_raw[idx] if idx < len(student_names_raw) else ""
                if sid or sname:
                    db.add(models.StudentMember(
                        session_id=session_id,
                        group_id=group.id,
                        program=program,
                        student_id=sid,
                        full_name=sname,
                    ))
        else:
            topics = [str(t).strip() for t in row[4:] if t is not None and str(t).strip()]
            group = models.Group(
                session_id=session_id,
                group_id=str(row[0]) if row[0] is not None else None,
                anonymous_code=str(row[1]) if row[1] is not None else None,
                program=program,
                representative=str(row[2]) if row[2] is not None else None,
                member_count=int(row[3]) if row[3] is not None else None,
                topic_interest=json.dumps(topics, ensure_ascii=False),
            )
            db.add(group)
            db.flush()
            if row[2] is not None and str(row[2]).strip():
                db.add(models.StudentMember(
                    session_id=session_id,
                    group_id=group.id,
                    program=program,
                    student_id=str(row[0]) if row[0] is not None else f"STD-{group.id}",
                    full_name=str(row[2]).strip(),
                ))


def _import_professor_info(ws, session_id: int, db: Session, program: Optional[str] = None):
    """
    ชีต Professor_Info รองรับทั้ง:
    - รูปแบบใหม่ (4 คอลัมน์): ProfID | ชื่อ-นามสกุลของอาจารย์ | ความเชี่ยวชาญ | โควต้ากลุ่ม
    - รูปแบบเดิม (5 คอลัมน์): ProfID | AnonymousCode | FullName | Expertise | Quota
    """
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    header_row = [str(c).strip().lower() if c else "" for c in (first_row or [])]
    is_new_format = len(header_row) == 4 or any("โควต้า" in h for h in header_row)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or str(row[0]).strip() == "":
            continue

        if is_new_format or len(row) == 4:
            prof_id = str(row[0]).strip()
            full_name = str(row[1]).strip() if len(row) > 1 and row[1] is not None else None
            expertise = str(row[2]).strip() if len(row) > 2 and row[2] is not None else None
            quota = int(row[3]) if len(row) > 3 and row[3] is not None else None

            prof = models.Professor(
                session_id=session_id,
                prof_id=prof_id,
                anonymous_code=prof_id,
                program=program,
                full_name=full_name,
                expertise=expertise,
                quota=quota,
            )
            db.add(prof)
        else:
            prof = models.Professor(
                session_id=session_id,
                prof_id=str(row[0]) if row[0] is not None else None,
                anonymous_code=str(row[1]) if row[1] is not None else None,
                program=program,
                full_name=str(row[2]) if row[2] is not None else None,
                expertise=str(row[3]) if row[3] is not None else None,
                quota=int(row[4]) if row[4] is not None else None,
            )
            db.add(prof)


def _import_student_rankings(ws, session_id: int, db: Session, program: Optional[str] = None):
    """
    ชีต Student_Rankings: GroupCode | ProfCode1 | ProfCode2 | ...
    Row 1 = header: GroupCode, then prof codes
    Row 2+ = group code, then ranks
    """
    prof_codes = [c.value for c in ws[1][1:] if c.value is not None]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        group_code = str(row[0])
        for i, pcode in enumerate(prof_codes):
            rank_val = row[i + 1]
            if rank_val is None:
                continue
            ranking = models.StudentRanking(
                session_id=session_id,
                program=program,
                group_code=group_code,
                prof_code=str(pcode),
                rank=int(rank_val),
            )
            db.add(ranking)


def _import_professor_scores(ws, session_id: int, db: Session, program: Optional[str] = None):
    """
    ชีต Professor_Scores: ProfCode | GroupCode | Score_A | Score_B | SubScore | MainScore
    """
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        pcode, gcode = str(row[0]), str(row[1])
        score_a = int(row[2]) if row[2] is not None else None
        score_b = int(row[3]) if row[3] is not None else None
        sub_score = float(row[4]) if row[4] is not None else None
        main_score = int(round(float(row[5]))) if row[5] is not None else None
        score = models.ProfessorScore(
            session_id=session_id,
            program=program,
            prof_code=pcode,
            group_code=gcode,
            score_a=score_a,
            score_b=score_b,
            sub_score=sub_score,
            main_score=main_score,
        )
        db.add(score)
